import sqlite3

from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QColorDialog, QFileDialog, QLabel, QScrollArea, QWidget, \
    QTableWidgetItem, QPushButton, QInputDialog
import sys

from yargy import rule, Parser
from yargy.predicates import gram, dictionary, normalized
from yargy.pipelines import morph_pipeline
from openpyxl import load_workbook



con = sqlite3.connect("test.db")


class MyWidget(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('1.ui', self)

        self.lineEdit.textChanged[str].connect(self.new_text)
        self.pushButton.clicked.connect(self.search_courses)

        self.position = ''
        self.lst_courses = []

        self.updata_tw()

    def new_text(self, text) -> None:
        self.position = text
        print(self.position)

    def search_courses(self):
        R1 = rule(gram('ADJF').optional(), gram('NOUN'))  # rule searching all key pairs ADJECTIVE NOUN
        parser = Parser(R1)

        comps = load_workbook('competences.xlsx')  # opening xlsx table as data source
        first = comps['Лист1']
        c = int(input())
        print(first[f'B{c}'].value)
        text = first[f'C{c}'].value

        self.res = set()
        for item in parser.findall(text):  # getting simple 'keywords' from table cell
            self.res.add(' '.join([_.value for _ in item.tokens]))
        print(self.res)

        R2 = morph_pipeline(list(self.res))  # rule searching matches by keywords in desc
        self.parser1 = Parser(R2)

        courses = load_workbook('courses.xlsx')  # loading page with courses' descriptions
        self.page = courses['Описание курсов']
        self.find_all_courses()
        self.updata_tw()

    def find_all_courses(self) -> set:
        list_of_matched = []  # resulting list
        for i in range(2, 500):
            list_of_matched.clear()  # cleaning list
            desc = self.page[f'B{i}'].value  # setting value of considered self.page

            for match in self.parser1.findall(desc):  # parsing
                list_of_matched.append(' '.join([_.value if len(_.value) > 1 else '' for _ in match.tokens]))

            if len(self.res.intersection(list_of_matched)) > len(
                    list_of_matched) // 10:  # checking if more than 20% of keywords are equal
                # print(list_of_matched)
                print(self.page[f'A{i}'].value)  # name of course

            # print(len(res.intersection(list_of_matched)) > len(res)//10)
            # print(page[f'A{i}'].value)  # name of course

        return set(list_of_matched)

    def updata_tw(self):
        self.con = sqlite3.connect("Сourses.db")
        cur = self.con.cursor()
        result = cur.execute("SELECT * FROM competencies").fetchall()
        print(result)
        self.tableWidget.setRowCount(len(result))
        self.tableWidget.setColumnCount(len(result[0]))
        for i, elem in enumerate(result):
            for j, val in enumerate(elem):
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(val)))
        # для ex
        '''
                self.result = [('', '')]
        # парс таблици
        self.result = self.result
        
        self.tableWidget.setRowCount(len(self.result))
        self.tableWidget.setColumnCount(len(self.result[0]))
        for i, elem in enumerate(self.result):
            for j, val in enumerate(elem):
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(val)))
        '''


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyWidget()
    ex.show()
    sys.exit(app.exec())
